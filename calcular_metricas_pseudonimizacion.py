"""
calcular_metricas_pseudonimizacion.py
=====================================
Calcula Precision, Recall y F1 del sistema de pseudonimización ARCOTEL PAS.

ENTRADAS:
    - fn_anotaciones.csv
    - vp_conteos.csv

SALIDA:
    - metricas_pseudonimizacion.xlsx
    - metricas_pseudonimizacion.txt
"""

import csv
from collections import defaultdict
from pathlib import Path

ENTIDAD_A_CAPA = {
    "RUC": "Capa 1 (regex)",
    "CEDULA": "Capa 1 (regex)",
    "EMAIL": "Capa 1 (regex)",
    "TELEFONO": "Capa 1 (regex) / 1.5 (contextual)",
    "DIRECCION": "Capa 1.5 (contextual)",
    "NOMBRE": "Capa 2 (spaCy NER)",
}

CAPA_NOMBRES = {
    "1_regex": "Capa 1 — Regex (datos estructurados)",
    "1.5_contextual": "Capa 1.5 — Contextual (encabezado)",
    "2_spacy": "Capa 2 — spaCy NER",
    "3_firmantes": "Capa 3 — Firmantes (regex)",
}


def deduplicar_csv(ruta: Path, subset_cols: list, descripcion: str) -> int:
    """
    Elimina filas duplicadas de un CSV in-place, manteniendo la última ocurrencia.
    Necesaria porque procesar_masivo_v2.ps1 hace append en cada ejecución.
    """
    if not ruta.exists():
        return 0

    with open(ruta, newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        # ✅ FIX: filtrar columnas None (coma extra al final del header)
        fieldnames = [fn for fn in reader.fieldnames if fn is not None]
        filas = []
        for row in reader:
            # ✅ FIX: eliminar claves None de cada fila
            fila_limpia = {k: v for k, v in row.items() if k is not None}
            filas.append(fila_limpia)

    total_original = len(filas)
    seen = {}
    for fila in filas:
        key = tuple(fila.get(col, '').strip() for col in subset_cols)
        seen[key] = fila

    filas_dedup = list(seen.values())
    eliminados = total_original - len(filas_dedup)

    with open(ruta, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(filas_dedup)

    if eliminados > 0:
        print(f"🧹 {descripcion}: {total_original} → {len(filas_dedup)} filas "
              f"({eliminados} duplicados eliminados)")
    else:
        print(f"✅ {descripcion}: {total_original} filas, sin duplicados")

    return eliminados


def cargar_fn(ruta_csv: str) -> list:
    fn_list = []
    with open(ruta_csv, newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row['resultado'].strip().upper() == 'FN':
                fn_list.append(row)
    return fn_list


def calcular_metricas(vp: int, fn: int, fp: int = 0) -> dict:
    precision = vp / (vp + fp) if (vp + fp) > 0 else 0.0
    recall = vp / (vp + fn) if (vp + fn) > 0 else 0.0
    f1 = (2 * precision * recall / (precision + recall)
          if (precision + recall) > 0 else 0.0)
    return {
        "VP": vp, "FN": fn, "FP": fp,
        "Total_real": vp + fn,
        "Precision": round(precision * 100, 1),
        "Recall": round(recall * 100, 1),
        "F1": round(f1 * 100, 1),
    }


def cargar_vp(ruta_csv: str) -> dict:
    vp = {}
    with open(ruta_csv, newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            doc = row['documento'].strip()
            vp[doc] = {
                'RUC': int(row.get('RUC', 0) or 0),
                'CEDULA': int(row.get('CEDULA', 0) or 0),
                'EMAIL': int(row.get('EMAIL', 0) or 0),
                'TELEFONO': int(row.get('TELEFONO', 0) or 0),
                'DIRECCION': int(row.get('DIRECCION', 0) or 0),
                'NOMBRE': int(row.get('NOMBRE', 0) or 0),
            }
    return vp


def main():
    fn_csv = Path("fn_anotaciones.csv")
    vp_csv = Path("vp_conteos.csv")

    for f in [fn_csv, vp_csv]:
        if not f.exists():
            print(f"❌ No se encontró {f}")
            return

    # ============================================================
    # PASO 0: DEDUPLICACIÓN AUTOMÁTICA
    # procesar_masivo_v2.ps1 hace append → puede haber duplicados
    # si se ejecutó más de una vez. Se limpian in-place aquí.
    # ============================================================
    print("\n🔍 Verificando duplicados en CSVs...")
    deduplicar_csv(vp_csv, ['documento'],
                   "vp_conteos.csv")
    deduplicar_csv(fn_csv, ['documento', 'entidad_valor', 'tipo_entidad'],
                   "fn_anotaciones.csv")
    print()

    # ============================================================
    # PASO 1: CARGAR DATOS
    # ============================================================
    fn_list = cargar_fn(str(fn_csv))
    VP_CONTEOS = cargar_vp(str(vp_csv))
    print(f"✅ FN cargados: {len(fn_list)}")
    print(f"✅ VP cargados: {len(VP_CONTEOS)} documentos")

    fn_por_tipo = defaultdict(int)
    fn_por_capa = defaultdict(int)
    fn_por_doc = defaultdict(list)
    fn_por_tipdoc = defaultdict(lambda: defaultdict(int))

    for fn in fn_list:
        tipo = fn['tipo_entidad'].strip().upper()
        capa = fn['capa'].strip()
        doc = fn['documento'].strip()
        tipo_doc = fn['tipo_doc'].strip()
        fn_por_tipo[tipo] += 1
        fn_por_capa[capa] += 1
        fn_por_doc[doc].append(fn)
        fn_por_tipdoc[tipo_doc][tipo] += 1

    vp_total = 0
    vp_por_tipo = defaultdict(int)
    vp_por_tipo_doc = defaultdict(lambda: defaultdict(int))

    for doc, conteos in VP_CONTEOS.items():
        tipo_doc = "informe" if "CTDG" in doc else "peticion"
        for tipo_ent, count in conteos.items():
            tipo_upper = tipo_ent.upper()
            vp_total += count
            vp_por_tipo[tipo_upper] += count
            vp_por_tipo_doc[tipo_doc][tipo_upper] += count

    fn_total = len(fn_list)
    docs_total = len(VP_CONTEOS)

    # ============================================================
    # TABLA 1: Métricas globales
    # ============================================================
    m_global = calcular_metricas(vp_total, fn_total)

    # ============================================================
    # TABLA 2: Métricas por componente
    # ============================================================
    capa_vp = {
        "1_regex": sum(vp_por_tipo.get(t, 0) for t in ["RUC", "CEDULA", "EMAIL", "TELEFONO"]),
        "1.5_contextual": vp_por_tipo.get("DIRECCION", 0),
        "2_spacy": vp_por_tipo.get("NOMBRE", 0),
        "3_firmantes": 0,
    }
    metricas_capa = {ck: calcular_metricas(capa_vp.get(ck, 0), fn_por_capa.get(ck, 0))
                     for ck in CAPA_NOMBRES}

    # ============================================================
    # TABLA 3: Métricas por tipo de entidad
    # ============================================================
    todos_tipos = set(list(vp_por_tipo.keys()) + list(fn_por_tipo.keys()))
    metricas_tipo = {t: calcular_metricas(vp_por_tipo.get(t, 0), fn_por_tipo.get(t, 0))
                     for t in sorted(todos_tipos)}

    # ============================================================
    # TABLA 4: Por documento
    # ============================================================
    docs_info = {}
    for doc, conteos in VP_CONTEOS.items():
        fn_doc = fn_por_doc.get(doc, [])
        fn_count = len(fn_doc)
        docs_info[doc] = {
            "tipo": "informe" if "CTDG" in doc else "petición",
            "VP": sum(conteos.values()),
            "FN": fn_count,
            "estado": "✅ Completo" if fn_count == 0 else f"⚠️  Parcial ({fn_count} FN)",
            "fn_detalle": [f['entidad_valor'] for f in fn_doc]
        }

    docs_completos = sum(1 for d in docs_info.values() if d['FN'] == 0)
    docs_parciales = docs_total - docs_completos

    # ============================================================
    # SALIDA TEXTO
    # ============================================================
    lineas = []
    sep = "=" * 65
    lineas.append(sep)
    lineas.append("  MÉTRICAS DE PSEUDONIMIZACIÓN — SISTEMA ARCOTEL PAS")
    lineas.append("  TFE — Iván | 2026")
    lineas.append(sep)

    lineas.append("\n📊 TABLA 1: MÉTRICAS GLOBALES DEL SISTEMA")
    lineas.append("-" * 65)
    lineas.append(f"  Entidades reales (VP+FN): {m_global['Total_real']}")
    lineas.append(f"  VP (pseudonimizados):     {m_global['VP']}")
    lineas.append(f"  FN (no pseudonimizados):  {m_global['FN']}")
    lineas.append(f"  FP (pseudonimización errónea): {m_global['FP']}")
    lineas.append(f"  Precision:  {m_global['Precision']}%")
    lineas.append(f"  Recall:     {m_global['Recall']}%")
    lineas.append(f"  F1-Score:   {m_global['F1']}%")

    lineas.append("\n📊 TABLA 2: MÉTRICAS POR COMPONENTE (CAPA)")
    lineas.append("-" * 65)
    lineas.append(f"  {'Componente':<35} {'VP':>4} {'FN':>4} {'P%':>6} {'R%':>6} {'F1%':>6}")
    lineas.append("  " + "-" * 60)
    for ck, nombre in CAPA_NOMBRES.items():
        m = metricas_capa[ck]
        n = nombre.split(" — ")[1] if " — " in nombre else nombre
        lineas.append(f"  {n:<35} {m['VP']:>4} {m['FN']:>4} "
                      f"{m['Precision']:>6.1f} {m['Recall']:>6.1f} {m['F1']:>6.1f}")

    lineas.append("\n📊 TABLA 3: MÉTRICAS POR TIPO DE ENTIDAD")
    lineas.append("-" * 65)
    lineas.append(f"  {'Tipo':<12} {'Total real':>10} {'VP':>4} {'FN':>4} "
                  f"{'P%':>6} {'R%':>6} {'F1%':>6}  Capa principal")
    lineas.append("  " + "-" * 60)
    for tipo, m in metricas_tipo.items():
        capa = ENTIDAD_A_CAPA.get(tipo, "—")
        lineas.append(f"  {tipo:<12} {m['Total_real']:>10} {m['VP']:>4} {m['FN']:>4} "
                      f"{m['Precision']:>6.1f} {m['Recall']:>6.1f} {m['F1']:>6.1f}  {capa}")

    lineas.append("\n📋 TABLA 4: RESUMEN POR DOCUMENTO")
    lineas.append("-" * 65)
    lineas.append(f"  Total documentos evaluados: {docs_total}")
    lineas.append(f"  Pseudonimización completa:  {docs_completos}/{docs_total} "
                  f"({docs_completos / docs_total * 100:.1f}%)")
    lineas.append(f"  Pseudonimización parcial:   {docs_parciales}/{docs_total} "
                  f"({docs_parciales / docs_total * 100:.1f}%)")
    lineas.append("")
    for doc, info in sorted(docs_info.items()):
        lineas.append(f"  {doc:<35} {info['estado']}")
        for fn_val in info['fn_detalle']:
            lineas.append(f"    └─ FN: {fn_val[:55]}")

    lineas.append("\n📌 NOTA METODOLÓGICA")
    lineas.append("-" * 65)
    lineas.append("  FP = 0 asumido. El sistema no pseudonimizó texto institucional")
    lineas.append("  (ARCOTEL, Ley Orgánica, etc.) gracias a la lista de excepciones.")
    lineas.append("  Los VP de NOMBRE (Capa 2) incluyen también nombres detectados")
    lineas.append("  por Capa 1.5 y Capa 3. Limitación metodológica documentada.")
    lineas.append(sep)

    output_txt = "\n".join(lineas)
    print(output_txt)
    with open("metricas_pseudonimizacion.txt", "w", encoding="utf-8") as f:
        f.write(output_txt)
    print("\n✅ Guardado: metricas_pseudonimizacion.txt")

    # ============================================================
    # EXPORTAR A EXCEL
    # ============================================================
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        hdr_font = Font(bold=True, color="FFFFFF")
        hdr_fill = PatternFill("solid", fgColor="2F4F8F")
        ok_fill = PatternFill("solid", fgColor="C6EFCE")
        warn_fill = PatternFill("solid", fgColor="FFEB9C")
        center = Alignment(horizontal="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        def hdr(ws, row, col, value, width=None):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = hdr_font;
            cell.fill = hdr_fill
            cell.alignment = center;
            cell.border = thin_border
            if width: ws.column_dimensions[get_column_letter(col)].width = width

        def dat(ws, row, col, value, fill=None, bold=False, center_align=False):
            cell = ws.cell(row=row, column=col, value=value)
            if fill: cell.fill = fill
            if bold: cell.font = Font(bold=True)
            if center_align: cell.alignment = center
            cell.border = thin_border
            return cell

        # Hoja 1: Global
        ws1 = wb.active;
        ws1.title = "1. Global"
        ws1.merge_cells("A1:C1")
        ws1["A1"] = "MÉTRICAS GLOBALES DEL SISTEMA DE PSEUDONIMIZACIÓN"
        ws1["A1"].font = Font(bold=True, size=12)
        for c, h in enumerate(["Métrica", "Valor", "Interpretación"], 1):
            hdr(ws1, 3, c, h, width=20)
        filas_g = [
            ("Entidades reales (VP+FN)", m_global['Total_real'], "Total de datos personales a pseudonimizar"),
            ("Verdaderos Positivos (VP)", m_global['VP'], "Pseudonimizados correctamente"),
            ("Falsos Negativos (FN)", m_global['FN'], "Datos que debían pseudonimizarse y no se hizo"),
            ("Falsos Positivos (FP)", m_global['FP'], "Texto no personal que fue pseudonimizado"),
            ("Precision", f"{m_global['Precision']}%", "De lo pseudonimizado, ¿cuánto era correcto?"),
            ("Recall", f"{m_global['Recall']}%", "De los datos reales, ¿cuántos se capturaron?"),
            ("F1-Score", f"{m_global['F1']}%", "Media armónica Precision/Recall"),
        ]
        for i, (m, v, interp) in enumerate(filas_g, 4):
            try:
                fill = ok_fill if i >= 8 and float(str(v).replace('%', '')) >= 80 else None
            except:
                fill = None
            dat(ws1, i, 1, m, bold=True);
            dat(ws1, i, 2, v, fill=fill, center_align=True)
            dat(ws1, i, 3, interp)

        # Hoja 2: Por Componente
        ws2 = wb.create_sheet("2. Por Componente")
        ws2.merge_cells("A1:G1")
        ws2["A1"] = "MÉTRICAS POR COMPONENTE (CAPA)";
        ws2["A1"].font = Font(bold=True, size=12)
        for c, h in enumerate(["Componente", "Técnica", "VP", "FN", "Precision %", "Recall %", "F1 %"], 1):
            hdr(ws2, 3, c, h, width=32 if c == 1 else 14)
        tecnicas = {"1_regex": "Determinística", "1.5_contextual": "Determinística",
                    "2_spacy": "IA (NER)", "3_firmantes": "Determinística"}
        for i, (ck, nombre) in enumerate(CAPA_NOMBRES.items(), 4):
            m = metricas_capa[ck]
            fill = ok_fill if m['F1'] >= 80 else (warn_fill if m['F1'] > 0 else None)
            dat(ws2, i, 1, nombre);
            dat(ws2, i, 2, tecnicas[ck], center_align=True)
            dat(ws2, i, 3, m['VP'], center_align=True)
            dat(ws2, i, 4, m['FN'], center_align=True, fill=warn_fill if m['FN'] > 0 else None)
            dat(ws2, i, 5, m['Precision'], center_align=True, fill=fill)
            dat(ws2, i, 6, m['Recall'], center_align=True, fill=fill)
            dat(ws2, i, 7, m['F1'], center_align=True, fill=fill)

        # Hoja 3: Por Tipo Entidad
        ws3 = wb.create_sheet("3. Por Tipo Entidad")
        ws3.merge_cells("A1:H1")
        ws3["A1"] = "MÉTRICAS POR TIPO DE ENTIDAD";
        ws3["A1"].font = Font(bold=True, size=12)
        for c, h in enumerate(["Tipo Entidad", "Total Real", "VP", "FN",
                               "Precision %", "Recall %", "F1 %", "Capa Principal"], 1):
            hdr(ws3, 3, c, h, width=18 if c in [1, 8] else 12)
        for i, (tipo, m) in enumerate(metricas_tipo.items(), 4):
            capa = ENTIDAD_A_CAPA.get(tipo, "—")
            fill = ok_fill if m['F1'] >= 80 else (warn_fill if m['F1'] > 0 else None)
            dat(ws3, i, 1, tipo, bold=True);
            dat(ws3, i, 2, m['Total_real'], center_align=True)
            dat(ws3, i, 3, m['VP'], center_align=True)
            dat(ws3, i, 4, m['FN'], center_align=True, fill=warn_fill if m['FN'] > 0 else None)
            dat(ws3, i, 5, m['Precision'], center_align=True, fill=fill)
            dat(ws3, i, 6, m['Recall'], center_align=True, fill=fill)
            dat(ws3, i, 7, m['F1'], center_align=True, fill=fill)
            dat(ws3, i, 8, capa)

        # Hoja 4: Por Documento
        ws4 = wb.create_sheet("4. Por Documento")
        ws4.merge_cells("A1:E1")
        ws4["A1"] = "RESUMEN POR DOCUMENTO (CONTEXTO EJECUTIVO)"
        ws4["A1"].font = Font(bold=True, size=12)
        for c, h in enumerate(["Documento", "Tipo", "VP", "FN", "Estado"], 1):
            hdr(ws4, 3, c, h, width=38 if c == 1 else 12)
        row = 4
        for doc, info in sorted(docs_info.items()):
            fill = ok_fill if info['FN'] == 0 else warn_fill
            dat(ws4, row, 1, doc);
            dat(ws4, row, 2, info['tipo'], center_align=True)
            dat(ws4, row, 3, info['VP'], center_align=True)
            dat(ws4, row, 4, info['FN'], center_align=True, fill=warn_fill if info['FN'] > 0 else None)
            dat(ws4, row, 5, "Completo" if info['FN'] == 0 else f"Parcial ({info['FN']} FN)",
                fill=fill, center_align=True)
            row += 1
            for fn_val in info['fn_detalle']:
                ws4.cell(row=row, column=1, value=f"  └─ FN: {fn_val[:60]}")
                ws4.cell(row=row, column=1).font = Font(italic=True, color="8B0000")
                row += 1
        row += 1
        ws4.cell(row=row, column=1, value="RESUMEN").font = Font(bold=True);
        row += 1
        ws4.cell(row=row, column=1, value=f"Documentos evaluados: {docs_total}");
        row += 1
        ws4.cell(row=row, column=1,
                 value=f"Pseudonimización completa: {docs_completos}/{docs_total} "
                       f"({docs_completos / docs_total * 100:.1f}%)");
        row += 1
        ws4.cell(row=row, column=1,
                 value=f"Pseudonimización parcial: {docs_parciales}/{docs_total} "
                       f"({docs_parciales / docs_total * 100:.1f}%)")

        wb.save("metricas_pseudonimizacion.xlsx")
        print("✅ Guardado: metricas_pseudonimizacion.xlsx")

    except ImportError:
        print("⚠️  openpyxl no instalado. Solo se generó el .txt.")
        print("   Instalar con: pip install openpyxl")


if __name__ == "__main__":
    main()
